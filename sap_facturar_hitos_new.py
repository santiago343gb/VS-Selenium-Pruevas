###########################################################################################
# sap_facturar_hitos_new.py — MULTIPROYECTO + EXCEL ÚNICO + OK/NOK EN MISMA HOJA
# Selección por FILTRO de cabecera (sin OCR) + clic por coordenadas (PyAutoGUI)
###########################################################################################

import os, time, json, re
import pandas as pd
from dotenv import load_dotenv
import openpyxl
import pyautogui as pag

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# ==============================
# CONFIG
# ==============================


EXCEL_PATH = r"C:\Users\bt00092\Downloads\tabla_facturar.xlsx"   # Excel único
RESULTADO_PATH = r"C:\Users\bt00092\Downloads\resultado_hitos.xlsx"  # reservado (no se usa)

CHROME_DRIVER_PATH = r"C:\Python Project\drivers\chromedriver.exe"

FAST_WAIT = 15
SLEEP_SHORT = 0.35
MAX_REINTENTOS = 3
RETRASO_ENTRE_REINTENTOS = 1.2

# fichero de calibración para coordenadas y pasos de menú
CALIB_MENU_FILE = "calibracion_hitos_menu.json"

# ==============================
# UTILIDADES BÁSICAS
# ==============================
def ensure_env():
    load_dotenv()
    u = os.getenv("FM21_USER2")
    p = os.getenv("FM21_PASS2")
    if not u or not p:
        raise Exception("Faltan credenciales en .env (FM21_USER2 / FM21_PASS2)")
    print(f"Usuario cargado: {u}")
    return u, p

def wait_no_busy(driver):
    try:
        WebDriverWait(driver, FAST_WAIT).until(
            EC.invisibility_of_element_located(
                (By.CSS_SELECTOR, ".sapUiBlockLayer, .sapUiLocalBusyIndicator")
            )
        )
    except:
        pass

def safe_type(driver, el, txt):
    try:
        el.clear()
        el.send_keys(txt)
    except:
        driver.execute_script("arguments[0].value='';", el)
        time.sleep(0.1)
        el.send_keys(txt)

def iniciar_driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-software-rasterizer")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--remote-debugging-pipe")
    opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    opts.add_experimental_option("useAutomationExtension", False)
    return webdriver.Chrome(service=Service(CHROME_DRIVER_PATH), options=opts)

# ==============================
# LOGIN
# ==============================
def login(driver, user, pwd):
    # ⚠️ importante: sin &amp; (el chat lo metía antes)
    URL = (
        "https://fm21global.tg.telefonica/fiori"
        "?sap-client=550&sap-language=ES"
        "#ZOBJ_Z_GESTION_HITOS_0001-display?sap-ie=edge&sap-theme=sap_belize"
    )
    driver.get(URL)
    time.sleep(1.2)
    driver.find_element(By.CSS_SELECTOR, "input[placeholder='Usuario']").send_keys(user)
    driver.find_element(By.CSS_SELECTOR, "input[placeholder='Clave de acceso']").send_keys(pwd)
    driver.find_element(By.XPATH, "//button[contains(text(),'Acceder')]").click()
    time.sleep(1.2)
    print("✔ Login OK")

# ==============================
# EJECUTAR PROYECTO
# ==============================
def ejecutar_proyecto(driver, proyecto):
    wait = WebDriverWait(driver, FAST_WAIT)

    wait.until(EC.frame_to_be_available_and_switch_to_it(
        (By.XPATH, "//iframe[contains(@id,'application-ZOBJ_Z_GESTION_HITOS')]")
    ))

    campo = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@title='Definición del proyecto']")
    ))
    safe_type(driver, campo, proyecto)
    time.sleep(SLEEP_SHORT)

    try:
        sug = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.XPATH, "//ul/li[1]"))
        )
        sug.click()
    except:
        try:
            campo.send_keys(Keys.ENTER)
        except:
            pass

    print("Buscando EJECUTAR…")
    try:
        btn = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.XPATH, "//*[normalize-space()='Ejecutar']/ancestor::button"))
        )
        driver.execute_script("arguments[0].click();", btn)
    except:
        ActionChains(driver).send_keys(Keys.F8).perform()

    driver.switch_to.default_content()
    wait_no_busy(driver)
    print("✔ Proyecto ejecutado correctamente")

# ==============================
# EXCEL — 1 SOLO ARCHIVO (lee y escribe Estado)
# ==============================
def cargar_excel():
    df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
    df["_row"] = df.index

    norm_cols = df.columns.str.lower().str.replace(" ", "", regex=False).str.replace(".", "", regex=False)
    colmap = dict(zip(norm_cols, df.columns))

    colp = next(c for c in norm_cols if ("proyecto" in c or "pep" in c))
    colh = next(c for c in norm_cols if ("hito" in c))

    df["proyecto_norm"] = df[colmap[colp]].astype(str).str.strip()
    df["hito_norm"] = df[colmap[colh]].astype(str).str.replace(".0", "", regex=False).str.strip()

    if "Estado" not in df.columns:
        df["Estado"] = ""

    return df, colmap[colp], colmap[colh]

def escribir_estado_en_excel(df, colp, colh, estados):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    if "Estado" not in headers:
        col_estado = ws.max_column + 1
        ws.cell(1, col_estado, "Estado")
    else:
        col_estado = headers["Estado"]

    for _, fila in df.iterrows():
        key = (fila["proyecto_norm"], fila["hito_norm"])
        if key in estados:
            estado = estados[key]
            fila_excel = fila["_row"] + 2
            ws.cell(row=fila_excel, column=col_estado, value=estado)

    wb.save(EXCEL_PATH)

# =============================================================================
# CALIBRACIÓN y MENÚ CONTEXTUAL (SIN OCR, SIN INSTALAR NADA)
# =============================================================================
def _load_calibr_menu():
    try:
        with open(CALIB_MENU_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

def _save_calibr_menu(cfg):
    with open(CALIB_MENU_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def calibracion_menu_guiada():
    """
    Pide 4 posiciones:
      1) X del checkbox (primera columna).
      2) X del checkbox FRD.
      3) X,Y de la cabecera de la columna Hito (para menú contextual).
      4) Y del centro de la primera fila de datos.
    Y dos números:
      - N_DOWN_FILTRO (flechas abajo hasta 'Filtrar…').
      - N_DOWN_LIMPIAR (flechas abajo hasta 'Quitar filtro/Restablecer').

    Guarda calibracion_hitos_menu.json
    """
    cfg = _load_calibr_menu()
    need = ("x_checkbox","x_frd","x_header_hito","y_header_hito","y_primera_fila","N_DOWN_FILTRO","N_DOWN_LIMPIAR")
    if all(k in cfg for k in need):
        return cfg

    pag.alert("CALIBRACIÓN (sin OCR)\n\n1) ALT+TAB a SAP.\n2) Sitúa el ratón EXACTO sobre un CHECKBOX de la 1ª columna.\n3) Vuelve aquí y pulsa OK.")
    x_cb, _ = pag.position()

    pag.alert("Ahora coloca el ratón EXACTO sobre un CHECKBOX de la columna FRD. Pulsa OK.")
    x_frd, _ = pag.position()

    pag.alert("Ahora coloca el ratón EXACTO sobre la CABECERA de la columna HITO. Pulsa OK.")
    x_hdr, y_hdr = pag.position()

    pag.alert("Ahora sitúa el ratón en el CENTRO de la 1ª FILA de datos (altura de fila). Pulsa OK.")
    _, y_row = pag.position()

    resp_fil = pag.prompt(text="¿Cuántas flechas ABAJO hasta 'Filtrar…' en el menú contextual?", default="2")
    resp_lim = pag.prompt(text="¿Cuántas flechas ABAJO hasta 'Quitar filtro/Restablecer'?", default="1")
    try: n_fil = int(resp_fil) if resp_fil is not None else 2
    except: n_fil = 2
    try: n_lim = int(resp_lim) if resp_lim is not None else 1
    except: n_lim = 1

    cfg = {
        "x_checkbox": int(x_cb),
        "x_frd": int(x_frd),
        "x_header_hito": int(x_hdr),
        "y_header_hito": int(y_hdr),
        "y_primera_fila": int(y_row),
        "N_DOWN_FILTRO": n_fil,
        "N_DOWN_LIMPIAR": n_lim,
    }
    _save_calibr_menu(cfg)
    pag.alert(f"Calibración guardada:\n{cfg}")
    return cfg

def _menu_contextual_ir_a_opcion(n_down: int):
    for _ in range(max(0, n_down)):
        pag.press("down"); time.sleep(0.05)
    pag.press("enter"); time.sleep(0.25)

def _abrir_filtrar_en_cabecera(cfg):
    pag.rightClick(cfg["x_header_hito"], cfg["y_header_hito"]); time.sleep(0.18)
    _menu_contextual_ir_a_opcion(cfg["N_DOWN_FILTRO"])

def _limpiar_filtro_cabecera(cfg):
    pag.rightClick(cfg["x_header_hito"], cfg["y_header_hito"]); time.sleep(0.18)
    _menu_contextual_ir_a_opcion(cfg["N_DOWN_LIMPIAR"])

# =============================================================================
# SELECCIÓN Y FRD por FILTRO (sin OCR)
# =============================================================================
def seleccionar_hitos_sin_ocr_via_filtro(driver, lista_hitos):
    """
    Para cada hito:
      - Abre 'Filtrar…' en cabecera Hito.
      - Escribe el número y ENTER (y ENTER extra por si aparece diálogo).
      - Clica checkbox de la 1ª fila (coordenada calibrada).
      - Limpia el filtro y pasa al siguiente.
    """
    WebDriverWait(driver, FAST_WAIT).until(
        EC.frame_to_be_available_and_switch_to_it(
            (By.ID, "application-ZOBJ_Z_GESTION_HITOS_0001-display-iframe")
        )
    )
    driver.switch_to.default_content()  # hacemos los clics en pantalla

    cfg = calibracion_menu_guiada()
    seleccionados = set()

    for hito in [str(h).strip() for h in lista_hitos]:
        print(f"\n🔎 Filtrando hito: {hito}")

        _abrir_filtrar_en_cabecera(cfg)
        pag.typewrite(hito, interval=0.02)
        pag.press("enter")
        time.sleep(0.6)
        # algunos ITS piden un Enter adicional para confirmar diálogo
        pag.press("enter"); time.sleep(0.2)

        try:
            pag.click(cfg["x_checkbox"], cfg["y_primera_fila"])
            time.sleep(0.12)
            seleccionados.add(hito)
            print(f"✔ Seleccionado: {hito}")
        except Exception as e:
            print(f"❌ No se pudo clicar checkbox para {hito}: {e}")

        _limpiar_filtro_cabecera(cfg)
        time.sleep(0.25)

    return seleccionados

def marcar_frd_sin_ocr_via_filtro(driver, lista_hitos):
    WebDriverWait(driver, FAST_WAIT).until(
        EC.frame_to_be_available_and_switch_to_it(
            (By.ID, "application-ZOBJ_Z_GESTION_HITOS_0001-display-iframe")
        )
    )
    driver.switch_to.default_content()

    cfg = calibracion_menu_guiada()
    marcados = set()

    for hito in [str(h).strip() for h in lista_hitos]:
        print(f"\n🔎 FRD para: {hito}")

        _abrir_filtrar_en_cabecera(cfg)
        pag.typewrite(hito, interval=0.02)
        pag.press("enter")
        time.sleep(0.6)
        pag.press("enter"); time.sleep(0.2)

        try:
            pag.click(cfg["x_frd"], cfg["y_primera_fila"])
            time.sleep(0.12)
            marcados.add(hito)
            print(f"✔ FRD marcado: {hito}")
        except Exception as e:
            print(f"❌ No se pudo clicar FRD para {hito}: {e}")

        _limpiar_filtro_cabecera(cfg)
        time.sleep(0.25)

    return marcados

# ==============================
# MODIFICACIÓN HITOS
# ==============================
def pulsar_modificacion_hitos(driver):
    driver.switch_to.default_content()
    wait_no_busy(driver)
    print("Pulsando Modificación Hitos…")
    for xp in [
        "//span[contains(text(),'Modificación Hitos')]/ancestor::div",
        "//div[contains(@id,'btn[25]')]",
    ]:
        try:
            btn = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(1)
            wait_no_busy(driver)
            print("✔ Modificación Hitos abierta")
            return
        except:
            pass

    ActionChains(driver).key_down(Keys.CONTROL).send_keys(Keys.F1).key_up(Keys.CONTROL).perform()
    print("✔ Modificación por Ctrl+F1")
    time.sleep(1)
    wait_no_busy(driver)

# ==============================
# GRABAR
# ==============================
def pulsar_grabar(driver):
    driver.switch_to.default_content()
    wait_no_busy(driver)
    print("Pulsando GRABAR…")
    for xp in [
        "//span[contains(text(),'Grabar')]/ancestor::div",
        "//div[contains(@id,'btn[11]')]",
    ]:
        try:
            btn = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(2)
            wait_no_busy(driver)
            print("✔ Grabado OK")
            return True
        except:
            pass

    try:
        ActionChains(driver).key_down(Keys.CONTROL).send_keys("s").key_up(Keys.CONTROL).perform()
        time.sleep(2)
        wait_no_busy(driver)
        print("✔ Grabado por Ctrl+S")
        return True
    except:
        print("❌ No se pudo pulsar Grabar")
        return False

# ==============================
# MAIN
# ==============================
def main():
    # Failsafe PyAutoGUI: ratón a esquina sup-izq aborta script
    pag.FAILSAFE = True

    user, pwd = ensure_env()
    df, colp, colh = cargar_excel()

    estados_global = {}

    for proyecto, grupo in df.groupby("proyecto_norm"):
        print("====================================")
        print("Procesando proyecto:", proyecto)
        print("====================================")

        # Saltar los que ya están OK
        grupo_a_facturar = grupo[grupo["Estado"].astype(str).str.upper() != "OK"]
        if grupo_a_facturar.empty:
            print(f"✔ Proyecto {proyecto} ya está totalmente OK — se salta.")
            continue

        hitos = [str(h).strip() for h in grupo_a_facturar["hito_norm"].tolist()]
        estado_por_hito = {h: "NOK" for h in hitos}

        for intento in range(1, MAX_REINTENTOS + 1):
            driver = None
            try:
                driver = iniciar_driver()
                login(driver, user, pwd)
                ejecutar_proyecto(driver, proyecto)

                # SELECCIÓN por filtro de cabecera
                seleccionados = seleccionar_hitos_sin_ocr_via_filtro(driver, hitos)

                # MODIFICACIÓN HITOS
                pulsar_modificacion_hitos(driver)

                # FRD por filtro de cabecera
                frd_marcados = marcar_frd_sin_ocr_via_filtro(driver, hitos)

                # GRABAR
                grabado_ok = pulsar_grabar(driver)

                # Consolidar resultado
                for h in hitos:
                    if (h in seleccionados) and (h in frd_marcados) and grabado_ok:
                        estado_por_hito[h] = "OK"
                    else:
                        estado_por_hito[h] = "NOK"

                print(f"✔ Proyecto {proyecto} completado en intento {intento}")
                break

            except Exception as e:
                print(f"❌ Error en intento {intento}: {e}")
            finally:
                if driver:
                    try:
                        driver.quit()
                    except:
                        pass
                time.sleep(RETRASO_ENTRE_REINTENTOS)

        for h in hitos:
            estados_global[(proyecto, h)] = estado_por_hito[h]

    escribir_estado_en_excel(df, colp, colh, estados_global)
    print("✔ PROCESO COMPLETO — Estados escritos en tabla_facturar.xlsx")

if __name__ == "__main__":
    main()