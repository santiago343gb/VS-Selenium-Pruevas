'''
=========================================================================================
# sap_facturar2026.py
# Fecha de creacion: 17/03/2026
# Correo: santiago.perezalbarran@telefonica.com
# Script para cambiar la fecha real de un hito de un proyecto en SAP
# Proceso que se realiza:
# 1. Iniciar sesión en SAP
# 2. Navegar a la transacción ZHITOS (https://fm21global.tg.telefonica/fiori?sap-client=550&sap-language=ES#ZOBJ_Z_GESTION_HITOS_0001-display?sap-ie=edge&sap-theme=sap_belize&sap-touch=0)
# 3. Buscar el proyecto y el hito
# 4. Cambiar la fecha real del hito (a traves de un click en el campo de fecha)
# 5. Guardar los cambios (otro click)
# 6. Cerrar sesión en SAP
=========================================================================================
'''
import os,sys, re,subprocess,json, codecs, time,shutil
import pandas as pd
import traceback
import openpyxl
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import ProcessPoolExecutor, as_completed
from contextlib import contextmanager
from datetime import datetime
from dotenv import load_dotenv
path_CURRENT=os.path.dirname(os.path.realpath(__file__))
sys.path.append(path_CURRENT+'./../')
from utilities.data import paths
from utilities.master import exportDF, totalTime, configureLogson

start=datetime.now()
load_dotenv()
# ==============================
# CONFIG
# ==============================
EXCEL_PATH = r"C:\Users\bt00092\Downloads\tabla_facturar.xlsx"
CHROME_DRIVER_PATH = r"C:\Python Project\drivers\chromedriver.exe"

FILTRO_NUM_HITO_YA_PREPARADO = False
FAST_WAIT = 12
SLEEP_SHORT = 0.2
SLEEP_MEDIUM = 0.5
SLEEP_LONG = 0.9

MAX_REINTENTOS = 2
RETRASO_ENTRE_REINTENTOS = 1


# ==============================
# UTILIDADES
# ==============================
def log(msg):
    print(msg, flush=True)


def ensure_env():
    load_dotenv()
    u = os.getenv("FM21_USER2")
    p = os.getenv("FM21_PASS2")
    if not u or not p:
        raise Exception("Faltan credenciales en .env (FM21_USER2 / FM21_PASS2)")
    log(f"Usuario cargado: {u}")
    return u, p


def wait_no_busy(driver):
    try:
        WebDriverWait(driver, FAST_WAIT).until(
            EC.invisibility_of_element_located(
                (By.CSS_SELECTOR, ".sapUiBlockLayer, .sapUiLocalBusyIndicator")
            )
        )
    except Exception:
        pass


def safe_type(driver, el, txt):
    txt = str(txt)
    try:
        el.click()
        time.sleep(0.2)
    except Exception:
        pass

    try:
        el.send_keys(Keys.CONTROL, "a")
        time.sleep(0.1)
        el.send_keys(Keys.BACKSPACE)
        time.sleep(0.1)
        el.send_keys(txt)
        return
    except Exception:
        pass

    try:
        el.clear()
        time.sleep(0.2)
        el.send_keys(txt)
        return
    except Exception:
        pass

    try:
        driver.execute_script("arguments[0].value='';", el)
        time.sleep(0.2)
        el.send_keys(txt)
    except Exception as e:
        raise Exception(f"No pude escribir en el campo: {e}")


def iniciar_driver():
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-software-rasterizer")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--remote-debugging-pipe")
    opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(service=Service(CHROME_DRIVER_PATH), options=opts)
    driver.set_page_load_timeout(60)
    return driver


def click_js(driver, el):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'center'});", el)
        time.sleep(0.2)
    except Exception:
        pass

    try:
        driver.execute_script("arguments[0].click();", el)
        return
    except Exception:
        pass

    try:
        el.click()
        return
    except Exception:
        pass

    ActionChains(driver).move_to_element(el).click().perform()


def switch_main_frame(driver):
    driver.switch_to.default_content()
    WebDriverWait(driver, FAST_WAIT).until(
        EC.frame_to_be_available_and_switch_to_it(
            (By.XPATH, "//iframe[contains(@id,'application-ZOBJ_Z_GESTION_HITOS')]")
        )
    )


def first_visible(driver, xpaths, timeout=6):
    if isinstance(xpaths, str):
        xpaths = [xpaths]

    end = time.time() + timeout
    while time.time() < end:
        for xp in xpaths:
            try:
                els = driver.find_elements(By.XPATH, xp)
                for e in els:
                    try:
                        if e.is_displayed():
                            return e
                    except Exception:
                        pass
            except Exception:
                pass
        time.sleep(0.2)
    return None


def all_visible(driver, xpaths, timeout=4):
    if isinstance(xpaths, str):
        xpaths = [xpaths]

    end = time.time() + timeout
    visibles = []
    while time.time() < end:
        visibles = []
        for xp in xpaths:
            try:
                els = driver.find_elements(By.XPATH, xp)
                for e in els:
                    try:
                        if e.is_displayed():
                            visibles.append(e)
                    except Exception:
                        pass
            except Exception:
                pass
        if visibles:
            return visibles
        time.sleep(0.2)
    return []


# ==============================
# LOGIN
# ==============================
def login(driver, user, pwd):
    URL = (
        "https://fm21global.tg.telefonica/fiori"
        "?sap-client=550&sap-language=ES"
        "#ZOBJ_Z_GESTION_HITOS_0001-display?sap-ie=edge&sap-theme=sap_belize"
    )
    driver.get(URL)
    time.sleep(2.5)

    driver.find_element(By.CSS_SELECTOR, "input[placeholder='Usuario']").send_keys(user)
    time.sleep(0.3)
    driver.find_element(By.CSS_SELECTOR, "input[placeholder='Clave de acceso']").send_keys(pwd)
    time.sleep(0.3)
    driver.find_element(By.XPATH, "//button[contains(text(),'Acceder')]").click()

    time.sleep(2.5)
    wait_no_busy(driver)
    log("✔ Login OK")


# ==============================
# EJECUTAR PROYECTO
# ==============================
def ejecutar_proyecto(driver, proyecto):
    wait = WebDriverWait(driver, FAST_WAIT)

    switch_main_frame(driver)

    campo = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@title='Definición del proyecto']")
    ))
    safe_type(driver, campo, proyecto)
    time.sleep(SLEEP_SHORT)

    try:
        sug = WebDriverWait(driver, 4).until(
            EC.element_to_be_clickable((By.XPATH, "//ul/li[1]"))
        )
        sug.click()
        time.sleep(0.8)
    except Exception:
        try:
            campo.send_keys(Keys.ENTER)
            time.sleep(0.8)
        except Exception:
            pass

    log("Buscando EJECUTAR…")
    try:
        btn = WebDriverWait(driver, 4).until(
            EC.element_to_be_clickable((By.XPATH, "//*[normalize-space()='Ejecutar']/ancestor::button"))
        )
        driver.execute_script("arguments[0].click();", btn)
    except Exception:
        ActionChains(driver).send_keys(Keys.F8).perform()

    driver.switch_to.default_content()
    time.sleep(1.5)
    wait_no_busy(driver)
    log("✔ Proyecto ejecutado correctamente")


# ==============================
# EXCEL
# ==============================
def cargar_excel():
    df = pd.read_excel(EXCEL_PATH, engine="openpyxl")
    df["_row"] = df.index

    norm_cols = df.columns.str.lower().str.replace(" ", "", regex=False).str.replace(".", "", regex=False)
    colmap = dict(zip(norm_cols, df.columns))

    colp = next(c for c in norm_cols if ("proyecto" in c or "pep" in c))
    colh = next(c for c in norm_cols if ("hito" in c))

    df["proyecto_norm"] = df[colmap[colp]].astype(str).str.strip()
    df["hito_norm"] = df[colmap[colh]].astype(str).str.replace(".0", "", regex=False).str.strip()

    if "Estado" not in df.columns:
        df["Estado"] = ""

    return df, colmap[colp], colmap[colh]


def escribir_estado_en_excel(df, colp, colh, estados):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

    if "Estado" not in headers:
        col_estado = ws.max_column + 1
        ws.cell(1, col_estado, "Estado")
    else:
        col_estado = headers["Estado"]

    for _, fila in df.iterrows():
        key = (fila["proyecto_norm"], fila["hito_norm"])
        if key in estados:
            estado = estados[key]
            fila_excel = fila["_row"] + 2
            ws.cell(row=fila_excel, column=col_estado, value=estado)

    wb.save(EXCEL_PATH)


# ==============================
# BOTONES SAP
# ==============================
def pulsar_modificacion_hitos(driver):
    driver.switch_to.default_content()
    wait_no_busy(driver)

    log("Comprobando si queda algún filtro abierto antes de Modificación Hitos…")
    cerrado = cerrar_dialogos_filtro_si_abiertos(driver)
    if not cerrado:
        raise Exception("Sigue abierto un diálogo de filtro y no pude cerrarlo.")

    driver.switch_to.default_content()
    wait_no_busy(driver)
    log("Pulsando Modificación Hitos…")

    for xp in [
        "//span[contains(text(),'Modificación Hitos')]/ancestor::*[self::button or self::div][1]",
        "//*[@title='Modificación Hitos']",
        "//*[@aria-label='Modificación Hitos']",
        "//div[contains(@id,'btn[25]')]",
    ]:
        try:
            btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, xp))
            )
            click_js(driver, btn)
            time.sleep(2.0)
            wait_no_busy(driver)
            log("✔ Modificación Hitos abierta")
            return True
        except Exception:
            pass

    try:
        ActionChains(driver).key_down(Keys.CONTROL).send_keys(Keys.F1).key_up(Keys.CONTROL).perform()
        time.sleep(2.0)
        wait_no_busy(driver)
        log("✔ Modificación por Ctrl+F1")
        return True
    except Exception:
        return False
def pulsar_marcar_fecha_real(driver):
    driver.switch_to.default_content()
    wait_no_busy(driver)
    log("Pulsando Marcar Fecha Real…")

    # SAP suele poner estos botones dentro del frame principal
    try:
        switch_main_frame(driver)
    except Exception:
        pass

    btn = first_visible(driver, [
        "//*[normalize-space()='Marcar Fecha Real']/ancestor::div[contains(@id,'btn')]",
        "//*[contains(@id,'btn') and .//*[normalize-space()='Marcar Fecha Real']]",
        "//span[contains(text(),'Marcar Fecha Real')]/ancestor::div[contains(@id,'btn')]",
        "//div[contains(@title,'Marcar Fecha Real')]"
    ], timeout=6)

    if not btn:
        raise Exception("No encontré el botón 'Marcar Fecha Real'.")

    click_js(driver, btn)
    time.sleep(2)
    wait_no_busy(driver)

    log("✔ Marcar Fecha Real pulsado")
    return True

def pulsar_grabar(driver):
    log("Pulsando GRABAR…")

    # En Modificación Hitos, mejor buscar directamente en el frame principal
    try:
        switch_main_frame(driver)
    except Exception:
        driver.switch_to.default_content()

    wait_no_busy(driver)

    btn = first_visible(driver, [
        "//*[normalize-space()='Grabar']/ancestor::div[contains(@id,'btn')]",
        "//*[contains(@id,'btn') and .//*[normalize-space()='Grabar']]",
        "//span[contains(text(),'Grabar')]/ancestor::div[contains(@id,'btn')]",
        "//div[contains(@title,'Grabar')]",
        "//div[contains(@aria-label,'Grabar')]"
    ], timeout=8)

    if btn:
        click_js(driver, btn)
        time.sleep(2.0)
        wait_no_busy(driver)
        log("✔ Grabado OK")
        return True

    # fallback por atajo de teclado si SAP no deja clicar bien el DOM
    try:
        ActionChains(driver).key_down(Keys.CONTROL).send_keys("g").key_up(Keys.CONTROL).perform()
        time.sleep(2.0)
        wait_no_busy(driver)
        log("✔ Grabado OK por atajo Ctrl+G")
        return True
    except Exception:
        pass

    raise Exception("No encontré el botón Grabar.")

# ==============================
# GRID / CELDAS / CHECKS
# ==============================
def encontrar_celda_base_para_filtrar(driver):
    switch_main_frame(driver)

    celda = first_visible(driver, [
        "//td[@role='gridcell']",
        "//div[@role='gridcell']",
        "//table//tr[1]//td[1]",
        "(//span[contains(@id,'grid#')])[1]"
    ], timeout=8)

    if not celda:
        raise Exception("No encontré una celda base para abrir el filtro.")
    return celda


def encontrar_celda_hito(driver, hito, timeout=8):
    switch_main_frame(driver)

    hito = str(hito).strip()

    xp_list = [
        f"//td[@role='gridcell'][normalize-space()='{hito}']",
        f"//div[@role='gridcell'][normalize-space()='{hito}']",
        f"//span[normalize-space()='{hito}']",
        f"//div[normalize-space()='{hito}']",
        f"//*[self::span or self::div or self::td][normalize-space()='{hito}']"
    ]

    celda = first_visible(driver, xp_list, timeout=timeout)
    if not celda:
        raise Exception(f"No encontré la celda del hito {hito}")
    return celda


def marcar_checkbox_seleccion(driver):
    switch_main_frame(driver)

    cb = first_visible(driver, [
        "//span[@role='checkbox' and @aria-checked='false']",
        "(//span[@role='checkbox'])[1]"
    ], timeout=6)

    if not cb:
        raise Exception("No encontré checkbox de selección.")
    click_js(driver, cb)
    time.sleep(0.6)
    wait_no_busy(driver)


def marcar_checkbox_frd(driver):
    switch_main_frame(driver)

    cb = first_visible(driver, [
        "//span[@role='checkbox' and @aria-checked='false']",
        "(//span[@role='checkbox'])[2]",
        "(//span[@role='checkbox'])[last()]"
    ], timeout=6)

    if not cb:
        raise Exception("No encontré checkbox FRD.")
    click_js(driver, cb)
    time.sleep(0.6)
    wait_no_busy(driver)


# ==============================
# FILTRO SAP
# ==============================
def abrir_menu_filtro(driver):
    switch_main_frame(driver)

    log("   -> Buscando botón filtro...")

    btn = first_visible(driver, [
        "//button[contains(@aria-label,'Filtro')]",
        "//button[contains(@title,'Filtro')]",
        "//*[@aria-label='Filtro']",
        "//*[@title='Filtro']",
        "//div[contains(@id,'MB_FILTER')]",
        "//div[contains(@id,'FILTER')]",
        "//span[contains(@title,'Filtro')]/ancestor::button[1]",
        "//span[contains(@aria-label,'Filtro')]/ancestor::button[1]"
    ], timeout=8)

    if not btn:
        raise Exception("No encontré el botón desplegable de filtro.")

    click_js(driver, btn)
    time.sleep(1.2)
    wait_no_busy(driver)
    log("   -> Botón filtro pulsado")
def buscar_opcion_menu_filtro(driver, texto_objetivo, timeout=5):
    """
    Busca la opción del menú tanto dentro del frame como fuera.
    """
    candidatos = [
        f"//*[normalize-space()='{texto_objetivo}']",
        f"//*[contains(normalize-space(),'{texto_objetivo}')]",
        f"//*[@title='{texto_objetivo}']",
        f"//*[@aria-label='{texto_objetivo}']",
    ]

    # primero fuera del frame
    driver.switch_to.default_content()
    opcion = first_visible(driver, candidatos, timeout=timeout)
    if opcion:
        return opcion, "default_content"

    # luego dentro del frame
    try:
        switch_main_frame(driver)
        opcion = first_visible(driver, candidatos, timeout=timeout)
        if opcion:
            return opcion, "frame"
    except Exception:
        pass

    return None, None


def elegir_opcion_menu_filtro(driver, texto_objetivo):
    log(f"   -> Buscando opción del menú: {texto_objetivo}")

    opcion, contexto = buscar_opcion_menu_filtro(driver, texto_objetivo, timeout=5)

    if not opcion:
        # debug extra: mirar si aparecen textos parecidos
        driver.switch_to.default_content()
        similares_default = all_visible(driver, [
            "//*[contains(normalize-space(),'Fijar')]",
            "//*[contains(normalize-space(),'filtro')]",
            "//*[contains(normalize-space(),'Borrar')]"
        ], timeout=1)

        try:
            switch_main_frame(driver)
            similares_frame = all_visible(driver, [
                "//*[contains(normalize-space(),'Fijar')]",
                "//*[contains(normalize-space(),'filtro')]",
                "//*[contains(normalize-space(),'Borrar')]"
            ], timeout=1)
        except Exception:
            similares_frame = []

        log(f"   -> Similares fuera del frame: {len(similares_default)}")
        log(f"   -> Similares dentro del frame: {len(similares_frame)}")

        raise Exception(f"No encontré la opción del menú: {texto_objetivo}")

    log(f"   -> Opción encontrada en: {contexto}")
    click_js(driver, opcion)
    time.sleep(1.0)
    wait_no_busy(driver)
    log(f"   -> Opción pulsada: {texto_objetivo}")


def esta_dialogo_filtro_abierto(driver, timeout=4):
    """
    Detecta el diálogo grande de filtro, tanto dentro del frame como fuera.
    """
    candidatos = [
        "//div[@role='dialog' and .//*[contains(normalize-space(),'Filtro')]]",
        "//div[contains(@class,'sapMDialog') and .//*[contains(normalize-space(),'Filtro')]]",
        "//*[contains(normalize-space(),'Paso 1: Definición de criterios filtrado')]",
        "//*[contains(normalize-space(),'Pool de columnas')]",
        "//*[contains(normalize-space(),'Criter.filtro')]",
        "//*[contains(normalize-space(),'Espec.valores')]"
    ]

    # fuera del frame
    driver.switch_to.default_content()
    dlg = first_visible(driver, candidatos, timeout=timeout)
    if dlg:
        return True, "default_content"

    # dentro del frame
    try:
        switch_main_frame(driver)
        dlg = first_visible(driver, candidatos, timeout=timeout)
        if dlg:
            return True, "frame"
    except Exception:
        pass

    return False, None


def cerrar_dialogos_filtro_si_abiertos(driver):
    """
    Cierra cualquier diálogo de filtro que haya quedado abierto
    antes de pasar a Modificación Hitos.
    """
    for _ in range(3):
        abierto, contexto = esta_dialogo_filtro_abierto(driver, timeout=1.5)
        if not abierto:
            return True

        if contexto == "default_content":
            driver.switch_to.default_content()
        else:
            try:
                switch_main_frame(driver)
            except Exception:
                driver.switch_to.default_content()

        btn_cerrar = first_visible(driver, [
            "//button[@title='Cerrar']",
            "//*[@aria-label='Cerrar']",
            "//*[normalize-space()='Cancelar']/ancestor::button[1]",
            "//button[.//*[normalize-space()='Cancelar']]",
            "//button[contains(@class,'sapMDialogCloseBtn')]",
            "//span[contains(@class,'sapMDialogCloseBtn')]/ancestor::button[1]"
        ], timeout=2)

        if btn_cerrar:
            click_js(driver, btn_cerrar)
            time.sleep(1.0)
            wait_no_busy(driver)
        else:
            try:
                ActionChains(driver).send_keys(Keys.ESCAPE).perform()
                time.sleep(1.0)
                wait_no_busy(driver)
            except Exception:
                pass

    abierto, _ = esta_dialogo_filtro_abierto(driver, timeout=1)
    return not abierto

def completar_popup_especificar_valores(driver, hito):
    wait_no_busy(driver)
    log("   -> Gestionando popup de valores...")

    popup = None
    driver.switch_to.default_content()
    popup = first_visible(driver, [
        "//div[@role='dialog' and .//*[contains(normalize-space(),'Especificar valores')]]",
        "//div[contains(@class,'sapMDialog') and .//*[contains(normalize-space(),'Especificar valores')]]",
        "//*[contains(normalize-space(),'Especificar valores p.criterios filtros')]"
    ], timeout=3)

    contexto = "default_content"
    if not popup:
        try:
            switch_main_frame(driver)
            popup = first_visible(driver, [
                "//div[@role='dialog' and .//*[contains(normalize-space(),'Especificar valores')]]",
                "//div[contains(@class,'sapMDialog') and .//*[contains(normalize-space(),'Especificar valores')]]",
                "//*[contains(normalize-space(),'Especificar valores p.criterios filtros')]"
            ], timeout=3)
            if popup:
                contexto = "frame"
        except Exception:
            pass

    if not popup:
        raise Exception("No encontré el popup 'Especificar valores'.")

    if contexto == "default_content":
        driver.switch_to.default_content()
    else:
        switch_main_frame(driver)

    campos = driver.find_elements(By.XPATH, "//input[not(@type='hidden') and not(@readonly)]")
    visibles = []
    for c in campos:
        try:
            if c.is_displayed() and c.is_enabled():
                visibles.append(c)
        except Exception:
            pass

    if not visibles:
        raise Exception("No encontré el input para el valor del hito.")

    inp = visibles[0]
    click_js(driver, inp)
    time.sleep(0.3)

    # borrar siempre valor anterior y escribir el nuevo
    try:
        inp.send_keys(Keys.CONTROL, "a")
        time.sleep(0.1)
        inp.send_keys(Keys.BACKSPACE)
        time.sleep(0.2)
    except Exception:
        pass

    safe_type(driver, inp, hito)
    time.sleep(0.8)
    wait_no_busy(driver)
    log(f"   -> Hito escrito: {hito}")

    inp.send_keys(Keys.ENTER)
    time.sleep(1.2)
    wait_no_busy(driver)
    log("   -> Popup de valores confirmado con Enter")

def reset_estado_filtro_proyecto():
    global FILTRO_NUM_HITO_YA_PREPARADO
    FILTRO_NUM_HITO_YA_PREPARADO = False

def completar_dialogo_filtro(driver, hito):
    """
    Primera vez en un proyecto:
    - flujo largo: Número de hito -> > -> Espec.valores

    Siguientes veces en el mismo proyecto:
    - flujo corto: ir directo a Espec.valores
    """
    global FILTRO_NUM_HITO_YA_PREPARADO

    wait_no_busy(driver)
    log("⚠ Se abrió el diálogo completo de Filtro. Lo gestiono...")

    abierto, contexto = esta_dialogo_filtro_abierto(driver, timeout=3)
    if not abierto:
        raise Exception("No detecté el diálogo de Filtro al intentar completarlo.")

    if contexto == "default_content":
        driver.switch_to.default_content()
    else:
        switch_main_frame(driver)

    if not FILTRO_NUM_HITO_YA_PREPARADO:
        log("   -> Primera vez del proyecto: preparo 'Número de hito'")

        num_hito = first_visible(driver, [
            "//div[contains(@class,'sapMList')]//*[normalize-space()='Número de hito']",
            "//div[contains(@class,'sapMList')]//*[normalize-space()='Numero de hito']",
            "//*[text()='Número de hito']",
            "//*[text()='Numero de hito']"
        ], timeout=8)

        if not num_hito:
            raise Exception("No encontré 'Número de hito' en el popup de filtro.")

        click_js(driver, num_hito)
        time.sleep(0.8)
        wait_no_busy(driver)
        log("   -> Seleccionado 'Número de hito'")

        btn_add = first_visible(driver, [
            "//*[@aria-label='Añadir criterio de filtro (F7)']",
            "//*[@title='Añadir criterio de filtro (F7)']",
            "//*[normalize-space()='>']/ancestor::button[1]",
            "//*[normalize-space()='>']"
        ], timeout=8)

        if not btn_add:
            raise Exception("No encontré el botón 'Añadir criterio de filtro (F7)'.")

        click_js(driver, btn_add)
        time.sleep(1.2)
        wait_no_busy(driver)
        log("   -> Pulsado botón añadir criterio")

        FILTRO_NUM_HITO_YA_PREPARADO = True
    else:
        log("   -> Filtro ya preparado en este proyecto, voy directo a Espec.valores")

    btn_espec = first_visible(driver, [
        "//*[@aria-label='Especificar valores filtro']",
        "//*[@title='Especificar valores filtro']",
        "//*[contains(normalize-space(),'Espec.valores')]"
    ], timeout=8)

    if not btn_espec:
        raise Exception("No encontré el botón 'Espec.valores'.")

    click_js(driver, btn_espec)
    time.sleep(1.2)
    wait_no_busy(driver)
    log("   -> Pulsado 'Espec.valores'")

    completar_popup_especificar_valores(driver, hito)

    time.sleep(1.2)
    wait_no_busy(driver)

    encontrar_celda_hito(driver, hito, timeout=6)
    log(f"✔ Filtro aplicado para hito {hito}")

def fijar_filtro_por_hito(driver, hito):
    celda = encontrar_celda_base_para_filtrar(driver)
    click_js(driver, celda)
    time.sleep(0.5)

    log(f"   -> Abriendo filtro para hito {hito}")
    abrir_menu_filtro(driver)

    log("   -> Pulsando 'Fijar filtro'")
    try:
        elegir_opcion_menu_filtro(driver, "Fijar filtro")
    except Exception:
        log("   -> Reintentando apertura del menú filtro...")
        abrir_menu_filtro(driver)
        elegir_opcion_menu_filtro(driver, "Fijar filtro")

    # esperar un poco más porque SAP tarda en pintar el popup
    time.sleep(2.0)
    wait_no_busy(driver)

    abierto, contexto = esta_dialogo_filtro_abierto(driver, timeout=4)
    if abierto:
        log(f"   -> Se abrió diálogo completo en: {contexto}")
        completar_dialogo_filtro(driver, hito)
    else:
        log("   -> SAP no mostró diálogo. Intento validar si el filtro rápido realmente filtró")

    # validación obligatoria
    try:
        encontrar_celda_hito(driver, hito, timeout=6)
        log(f"   -> Hito {hito} visible tras filtro")
    except Exception:
        raise Exception(f"El filtro no dejó visible el hito {hito}")


def borrar_filtros(driver):
    try:
        abrir_menu_filtro(driver)
        try:
            elegir_opcion_menu_filtro(driver, "Borrar filtros")
        except Exception:
            log("   -> Reintentando menú para borrar filtros...")
            abrir_menu_filtro(driver)
            elegir_opcion_menu_filtro(driver, "Borrar filtros")

        time.sleep(1.2)
        wait_no_busy(driver)
        log("   -> Filtros borrados")
    except Exception as e:
        raise Exception(f"No pude borrar filtros: {e}")


# ==============================
# SELECCIÓN DE HITOS
# ==============================
def seleccionar_hitos(driver, lista_hitos):
    seleccionados = set()

    for hito in [str(h).strip() for h in lista_hitos]:
        try:
            log(f"\n🔎 Seleccionando hito: {hito}")

            fijar_filtro_por_hito(driver, hito)

            # validar que el hito filtrado está visible
            encontrar_celda_hito(driver, hito, timeout=5)

            # marcar el checkbox del hito visible
            marcar_checkbox_seleccion(driver)

            seleccionados.add(hito)
            log(f"✔ Seleccionado: {hito}")

        except Exception as e:
            log(f"❌ Error seleccionando hito {hito}: {e}")
            try:
               cerrar_dialogos_filtro_si_abiertos(driver)
            except Exception:
                pass

    # al terminar, asegurarnos de que no queda el filtro abierto
            try:
               cerrar_dialogos_filtro_si_abiertos(driver)
            except Exception:
               pass

    return seleccionados

# ==============================
# MAIN
# ==============================
def main():
    user, pwd = ensure_env()
    df, colp, colh = cargar_excel()

    estados_global = {}

    for proyecto, grupo in df.groupby("proyecto_norm"):
        log("====================================")
        log(f"Procesando proyecto: {proyecto}")
        log("====================================")
#reset valores filtro
        reset_estado_filtro_proyecto()

        grupo_a_facturar = grupo[grupo["Estado"].astype(str).str.upper() != "OK"]
        if grupo_a_facturar.empty:
            log(f"✔ Proyecto {proyecto} ya está totalmente OK — se salta.")
            continue

        hitos = [str(h).strip() for h in grupo_a_facturar["hito_norm"].tolist()]
        estado_por_hito = {h: "NOK" for h in hitos}

        for intento in range(1, MAX_REINTENTOS + 1):
            driver = None
            try:
                driver = iniciar_driver()
                login(driver, user, pwd)
                ejecutar_proyecto(driver, proyecto)

                seleccionados = seleccionar_hitos(driver, hitos)

                if not seleccionados:
                    raise Exception("No se seleccionó ningún hito. Se aborta el proyecto.")

                hitos_para_modificar = [h for h in hitos if h in seleccionados]
                if not hitos_para_modificar:
                    raise Exception("No hay hitos seleccionados para Modificación Hitos.")

                ok_mod = pulsar_modificacion_hitos(driver)
                if not ok_mod:
                    raise Exception("No se pudo abrir Modificación Hitos")

                ok_fecha_real = pulsar_marcar_fecha_real(driver)
                grabado_ok = pulsar_grabar(driver)

                if ok_fecha_real and grabado_ok:
                   for h in hitos:
                     estado_por_hito[h] = "OK"
                   else:
                    for h in hitos:
                       estado_por_hito[h] = "NOK"

                log(f"✔ Proyecto {proyecto} completado en intento {intento}")
                break

            except Exception as e:
                log(f"❌ Error en intento {intento}: {e}")
                log(traceback.format_exc())

            finally:
                if driver:
                    try:
                        driver.quit()
                    except Exception:
                        pass
                time.sleep(RETRASO_ENTRE_REINTENTOS)

        for h in hitos:
            estados_global[(proyecto, h)] = estado_por_hito[h]

    escribir_estado_en_excel(df, colp, colh, estados_global)
    log("✔ PROCESO COMPLETO — Estados escritos en tabla_facturar.xlsx")

def timexHito(start, hitos_len):
    if hitos_len == 0: return 0
    tiempo_total = (datetime.now() - start).total_seconds()
    tiempo_por_hito = tiempo_total / hitos_len
    return tiempo_por_hito




#=====================================================================
#=====================================================================
if __name__ == "__main__":

    mylogs = configureLogson(__name__,paths['logs_local']+'santi/'+os.path.basename(__file__).rsplit('.', 1)[0]+".log")
    mylogs.info("======================================================")
    mylogs.info("[START-TIME]["+str(start)+"]")

    mylogs.info("Script: Estamos probando el script de santiago")
    mylogs.error("Esto es un ERROR de script")

    # EJECUTAR SCRIPT PRINCIPAL
    main()

    mylogs.info("[END-TIME]["+str(datetime.now())+"]")
    mylogs.info('Total runtime: '+ str(totalTime(start, datetime.now())[0])+ ' minutes '+ str("{:.3f}".format(totalTime(start, datetime.now())[1]))+ ' seconds')
    mylogs.info("======================================================")